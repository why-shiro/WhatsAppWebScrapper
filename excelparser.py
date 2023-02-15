import os
import time

import openpyxl


class Excel:

    def __init__(self):
        if not(os.path.exists("contacts.xlsx")):
            wb = openpyxl.Workbook()
            time.sleep(1)
            wb.save("contacts.xlsx")
            time.sleep(5)
            self.workbook = openpyxl.load_workbook("contacts.xlsx")
            self.worksheet = self.workbook.active
            self.column = 1
            print("Dosya oluşturma başarılı")
        else:
            while True:
                print("Eğer baştan tarama yaparsanız tüm dosya silinir ve sıfırdan kaydedilir.")
                print(
                    "Eğer baştan tarama yapmazsanız giriş yapıldıktan sonra en üstten tarama işlemine başlanır ve listeye eklenir.")
                inp = str(input("Numaraları baştan taramak ister misiniz (Y/H)"))
                if inp == "Y":
                    print("Dosya temizleniyor, sıfırdan tarama işlemi iin hazırlanılıyor...")
                    time.sleep(1)
                    try:
                        os.remove("contacts.xlsx")
                        time.sleep(1)
                        wb = openpyxl.Workbook()
                        time.sleep(1)
                        wb.save("contacts.xlsx")
                        self.workbook = openpyxl.load_workbook("contacts.xlsx")
                        self.worksheet = self.workbook.active
                        print("Dosya oluşturma başarılı")
                        self.column = 1
                        break
                    except:
                        print("Dosya silme işlemi başarısız!")
                        exit()
                else:
                    print("En sondan yazdırma işlemine devam ediliyor... Tarama başlatılıyor.")
                    print("Eğer bu işlemi istemiyorsanız lütfen programı kapatın. (5s)")
                    time.sleep(5)
                    self.workbook = openpyxl.load_workbook("contacts.xlsx")
                    self.worksheet = self.workbook.active
                    self.max = len(self.worksheet['A'])
                    self.column = self.max + 1
                    print("Listed Numbers: ", self.max)
                    break


    def addContact(self, element):
        self.worksheet["A"+str(self.column)] = element
        self.workbook.save("contacts.xlsx")
        self.column += 1

ep = Excel()

